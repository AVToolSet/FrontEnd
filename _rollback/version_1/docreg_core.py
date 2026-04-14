from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

DRAWING_TYPES: list[tuple[int, str]] = [
    (0, "Drawing Schedule"),
    (1, "Issue Log"),
    (2, "Change Log"),
    (10, "Combined M&E"),
    (11, "Elevation"),
    (12, "Plan"),
    (13, "High Level"),
    (14, "Low Level"),
    (15, "Key Plan"),
    (16, "Room Booking Plan"),
    (17, "Custom Metalwork"),
    (18, "Furniture Design"),
    (19, "Equipment Detail"),
    (20, "Lighting Combined"),
    (21, "Lighting"),
    (22, "Load Schedule"),
    (23, "Lighting Terminals Plan"),
    (24, "Panel Layout"),
    (30, "System Schematic"),
    (31, "Video Schematic"),
    (32, "Audio Schematic"),
    (33, "Control Schematic"),
    (34, "Power Schematic"),
    (35, "Rack Layout"),
    (36, "I/O Panel Layout"),
]

DRAWING_NAME_BY_CODE = dict(DRAWING_TYPES)
GREY_FILL = PatternFill("solid", fgColor="C8C8C8")
ARIAL_12 = Font(name="Arial", size=12)
ARIAL_12_BOLD = Font(name="Arial", size=12, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass(slots=True)
class AreaRequirement:
    name: str
    description: str = ""
    selected_codes: set[int] | None = None

    def __post_init__(self) -> None:
        if self.selected_codes is None:
            self.selected_codes = set()

    @property
    def base_name(self) -> str:
        if "(" in self.name:
            return self.name.split("(", 1)[0].strip()
        return self.name.strip()


def validate_seqf(seqf_number: str) -> str:
    value = seqf_number.strip()
    if len(value) != 6 or not value.isdigit():
        raise ValueError("SEQF Number must be exactly 6 digits.")
    return value


def get_document_type_from_drawing_name(drawing_name: str) -> str:
    nm = drawing_name.strip().upper()

    if nm in {"DRAWING SCHEDULE", "ISSUE LOG", "CHANGE LOG"}:
        return "Schedule"

    if nm in {
        "COMBINED M&E",
        "ELEVATION",
        "PLAN",
        "HIGH LEVEL",
        "LOW LEVEL",
        "KEY PLAN",
        "ROOM BOOKING PLAN",
        "CUSTOM METALWORK",
        "FURNITURE DESIGN",
        "EQUIPMENT DETAIL",
    }:
        return "2D Drawing - AV M&E"

    if nm in {
        "LIGHTING COMBINED",
        "LIGHTING",
        "LOAD SCHEDULE",
        "LIGHTING TERMINALS PLAN",
        "PANEL LAYOUT",
        "SYSTEM SCHEMATIC",
        "VIDEO SCHEMATIC",
        "AUDIO SCHEMATIC",
        "CONTROL SCHEMATIC",
        "POWER SCHEMATIC",
        "RACK LAYOUT",
        "I/O PANEL LAYOUT",
    }:
        return "2D Drawing - AV Schematic"

    return "Unknown"


def clean_for_reg_subtype(doc_sub_type: str) -> str:
    return doc_sub_type.strip().replace("2D Drawing - AV ", "").replace("2D Drawing - ", "").strip()


def generate_document_register_rows(seqf_number: str, areas: list[AreaRequirement]) -> list[list[str]]:
    seqf = validate_seqf(seqf_number)
    base_area_order: dict[str, int] = {}
    for area in areas:
        if area.base_name and area.base_name not in base_area_order:
            base_area_order[area.base_name] = len(base_area_order) + 1

    area_drawing_counts: dict[tuple[str, int], int] = defaultdict(int)
    rows: list[list[str]] = []

    for area in areas:
        if not area.name.strip():
            continue

        folder_num = base_area_order[area.base_name]
        for drawing_type in sorted(area.selected_codes or set()):
            drawing_name = DRAWING_NAME_BY_CODE.get(drawing_type, "Unknown Drawing Type")
            clean_sub_type = clean_for_reg_subtype(get_document_type_from_drawing_name(drawing_name))
            area_key = (area.base_name, drawing_type)
            area_drawing_counts[area_key] += 1

            folder_code = f"{folder_num:02d}"
            type_code = f"{drawing_type:02d}"
            seq_code = f"{area_drawing_counts[area_key]:02d}"
            drawing_number = f"{seqf}-PAV-ZZ-ZZ-DR-X-{folder_code}{type_code}{seq_code}"

            rows.append([drawing_number, "2D Drawing", clean_sub_type, f"{area.name} - {drawing_name}"])

    return rows


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max(length + 2, 12)


def style_range(worksheet, cell_range: str, header_range: str) -> None:
    for row in worksheet[cell_range]:
        for cell in row:
            cell.font = ARIAL_12
            cell.border = THIN_BORDER

    for row in worksheet[header_range]:
        for cell in row:
            cell.font = ARIAL_12_BOLD
            cell.fill = GREY_FILL
            cell.border = THIN_BORDER


def build_workbook(seqf_number: str, areas: list[AreaRequirement]) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Document Register"
    worksheet.append(["Document Number", "Document Type", "Document Sub Type", "Document Name"])
    for row in generate_document_register_rows(seqf_number, areas):
        worksheet.append(row)
    style_range(worksheet, f"A1:D{max(worksheet.max_row, 2)}", "A1:D1")
    autosize_worksheet(worksheet)
    return workbook


def save_workbook(output_path: str | Path, seqf_number: str, areas: list[AreaRequirement]) -> Path:
    path = Path(output_path)
    build_workbook(seqf_number, areas).save(path)
    return path


def workbook_bytes(seqf_number: str, areas: list[AreaRequirement]) -> bytes:
    buffer = BytesIO()
    build_workbook(seqf_number, areas).save(buffer)
    return buffer.getvalue()
